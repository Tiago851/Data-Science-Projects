"""
This is a project I developed for my current employer - Exide Technologies SSC - whose main objective is to produce a detailed
remittance PDF file with all the documents paid on a certain date. 

In the first place, the script tests if the source data Excel file is present in the same folder as the script. In case the file is not
present a message will pop up and warn the user, otherwise it will run the main script.

- update
The supplier_vmf() searches each supplier's Master Data in from an independent tab that must be updated regularly as new suppliers
are created. This information is pasted on the template and if a supplier's data is not found, it warns the user.

Due to confidentiality reasons, the source file is no provided but an example of the output is available and named: 

-Payment_remittances_v1_output_example

"""

#Modules
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import xlwings as xw
import os
from tkinter import *
from tkinter import messagebox

#Redirecting the working directory to our current working directory
abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)

#Shutting down all Excel applications to make sure the script runs correctly
os.system('taskkill /F /IM excel.exe')

#Confirming if the source file is in the same folder
def find():
    for root, dirs, files in os.walk(os.getcwd()):
        if 'teste_dfs.xlsx' in files:
            return True
        else:
            root = Tk()
            root.withdraw()
            root.geometry("300x200")

            w = Label(root, text ='teste', font = "50") 
            w.pack()
            messagebox.showwarning('File not found!', 'The Excel file for remittances generation is not in this folder. Please put the Excel file in the same directory as the executable and run the program again.')

            root.destroy()

            return False

#VMF function to search and find each supplier's data
def supplier_vmf(supplier):
    
    #The original dataframe
    df_original_vmf = pd.read_excel('teste_dfs.xlsx', sheet_name = 2)
    
    #Removing spaces from supplier codes
    df_original_vmf['CFOU'] = df_original_vmf['CFOU'].apply(lambda x: x.replace(" ", ""))
    
    #Creating a new dataframe to make reading the data easier and lighter
    df_vmf = pd.DataFrame(columns = ['Supplier Code', 'Address', 'Zip-code', 'City', 'VAT', 'Country'])
    
    #Assigning the data to the new dataframe
    df_vmf['Supplier Code'] = df_original_vmf['CFOU']
    df_vmf['Address'] = df_original_vmf['CADR1']
    df_vmf['Zip-code'] = df_original_vmf['SOPO']
    df_vmf['City'] = df_original_vmf['SODE']
    df_vmf['VAT'] = df_original_vmf['CNUMTVANEW']
    df_vmf['Country'] = df_original_vmf['LPAYS']
    
    #The supplier data
    supp_vmf = df_vmf.loc[df_vmf['Supplier Code'] == supplier ]
    
    #Output
    supp_vmf_list = []
    
    #Exception handling - it is necessary to inform the user that the supplier has no data
    try:
        address = supp_vmf['Address'].iloc[0].rstrip()
        zip_code = supp_vmf['Zip-code'].iloc[0].replace(" ", "")
        city = supp_vmf['City'].iloc[0].rstrip()
        country = supp_vmf['Country'].iloc[0].rstrip()
        vat = supp_vmf['VAT'].iloc[0].replace(" ", "")
    
    except IndexError:
        root = Tk()
        root.withdraw()
        root.geometry("300x200")

        w = Label(root, text ='teste', font = "50") 
        w.pack()
        
        messagebox.showwarning('Supplier not Valid', f'Supplier {supplier} is not on the VMF table. Please update the VMF data.')

        root.destroy()
        
        address = ""
        zip_code = ""
        city = ""
        country = ""
        vat = ""
        
    finally:
        supp_vmf_list.append(address)
        supp_vmf_list.append(zip_code)
        supp_vmf_list.append(city)
        supp_vmf_list.append(country)
        supp_vmf_list.append(vat)
    
    return supp_vmf_list


def payment_generator():
    #Reading the Excel file
    df_original = pd.read_excel('teste_dfs.xlsx', sheet_name = 0)

    #Renaming the dataframe for better reading
    df_final = pd.DataFrame(columns = ['Supplier Code', 
                                    'Supplier Name',
                                    'Invoice Journal'
                                    'Invoice Date',
                                    'EASY Number',
                                    'Amount',
                                    'Currency'])

    #Assigning the data
    df_final['Supplier Code'] = df_original['P5CFOU']
    df_final['Invoice Journal'] = df_original['P5CJL']
    df_final['Supplier Name'] = df_original['LFOU']
    df_final['Invoice Date'] = df_original['P5DPIE']
    df_final['EASY Number'] = df_original['P5CPIE']
    df_final['Amount'] = df_original['P5MRGL']
    df_final['Currency'] = df_original['P5CDVB']

    #List of unique supplier codes
    list_of_suppliers = list(set(df_final['Supplier Code']))

    #Looping through the list of suppliers
    for supplier in list_of_suppliers[:2]:
        
        #Opening the Workbook
        wb = openpyxl.load_workbook('teste_dfs.xlsx')
        
        template = wb['Template']
        
        one_sup = df_final.loc[df_final['Supplier Code'] == supplier ]

        #Executing the VMF function
        vmf_data = supplier_vmf(supplier)

        #Supplier Code
        template.cell(1,2).value = one_sup['Supplier Code'].iloc[0]
        
        #Supplier name
        template.cell(2,2).value = one_sup['Supplier Name'].iloc[0]

        #Supplier address
        template.cell(3,2).value = vmf_data[0]

        #Zip code + city
        zip_code = f'{vmf_data[1]}, {vmf_data[2]}'
        template.cell(4,2).value = zip_code

        #Country
        template.cell(5,2).value = vmf_data[3] 

        #VAT 
        template.cell(6,2).value = vmf_data[4]

        for i in range(0,len(one_sup)):

            #Invoice Journal
            template.cell(13+i,2).value = one_sup['Invoice Journal'].iloc[i]
            template.cell(13+i,2).alignment = Alignment(horizontal='center')
            template.cell(13+i,2).font = Font(name = 'Ariel', size = 11)

            #EASY number passing as invoice number
            template.cell(13+i,3).value = one_sup['EASY Number'].iloc[i]
            template.cell(13+i,3).alignment = Alignment(horizontal='center')
            template.cell(13+i,3).font = Font(name = 'Ariel', size = 11)

            #Invoice date
            inv_date = str(one_sup['Invoice Date'].iloc[i])
            inv_date = inv_date[6:] + '-' + inv_date[4:6] + '-' + inv_date[:4]
            template.cell(13+i,4).value = inv_date
            template.cell(13+i,4).alignment = Alignment(horizontal='center')
            template.cell(13+i,4).font = Font(name = 'Ariel', size = 11)

            #Amount
            template.cell(13+i,5).value = one_sup['Amount'].iloc[i]
            template.cell(13+i,5).alignment = Alignment(horizontal='center')
            template.cell(13+i,5).font = Font(name = 'Ariel', size = 11)

            #Currency
            template.cell(13+i,6).value = one_sup['Currency'].iloc[i]
            template.cell(13+i,6).alignment = Alignment(horizontal='center')
            template.cell(13+i,6).font = Font(name = 'Ariel', size = 11)

            #Total
            total_text = template.cell(14+len(one_sup),4)
            total_text.value = 'Total'
            total_text.font = Font(name = 'Ariel', size = 12, bold = True)
            total_text.alignment = Alignment(horizontal='center')
            template.cell(14+len(one_sup),5).value = one_sup['Amount'].sum()
            template.cell(14+len(one_sup),5).font = Font(name = 'Ariel', size = 11)
            template.cell(14+len(one_sup),5).alignment = Alignment(horizontal='center')
        
        #File name concatenation
        files = supplier + '.xlsx'
        
        #Saving & closing the file, otherwise it won't print correctly
        """
        There is probably a better & more efficient way to do this but this way works too...
        """

        wb.save(files)
        wb.close()
        
        app = xw.App(visible = False)
        
        #The error handling here is necessary as the xw.Book may file if the source file is transferred to the same folder
        #as the script
        try: 
            book = xw.Book(fullname=files).sheets[1]
        except OSError:
            os.system('taskkill /F /IM excel.exe')
        finally:
            book = xw.Book(fullname=files).sheets[1]

        path_to_pdf = f'{dname}\\{supplier}.pdf'
        
        book.api.ExportAsFixedFormat(0, path_to_pdf)
        
        app.quit()
        
        #Clearing the previous supplier data
        wb = openpyxl.load_workbook('teste_dfs.xlsx')
        
        wb['Template'].cell(2,2).value = None
        wb['Template'].cell(1,2).value = None
        
        for row in wb['Template']['B13:F2000']:
            for cell in row:
                cell.value = None
        
        #Saving the cleared file 
        wb.save('teste_dfs.xlsx')
        wb.close()

    #Deleting the excel files that the script creates
    try:
        for dirpath, dirname, filename  in os.walk(os.getcwd()):
            for ficheiro in filename:
                if ficheiro[-5:] == '.xlsx' and ficheiro != 'teste_dfs.xlsx':
                    os.remove(ficheiro)
    except IOError:
        openpyxl.Workbook(ficheiro).close()

    finally:
        for dirpath, dirname, filename  in os.walk(os.getcwd()):
            for ficheiro in filename:
                if ficheiro[-5:] == '.xlsx' and ficheiro != 'teste_dfs.xlsx':
                    os.remove(ficheiro)

if __name__ == '__main__' and find():
    #Executing the functions
    payment_generator()